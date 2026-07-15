#!/usr/bin/env python3
"""Convert a shipment Excel workbook to the lightweight data consumed by the
Container Explorer static site.

Only the 16 columns needed by the UI are kept. Header matching is
case-/space-/punctuation-insensitive so minor source-file drift is tolerated.

To stay safe on very large workbooks (hundreds of thousands of rows) the JSON is
written *streaming* (one record at a time) instead of building a giant list, and
an optional --gzip produces a compressed data.json.gz that is far smaller to
host on GitHub Pages and faster to download.

Usage:
    python scripts/convert.py input.xlsx [--out data.json] [--meta meta.json] [--gzip]
"""
import argparse
import gzip
import json
import re
import sys
from datetime import datetime, timezone

# Column -> acceptable header variants (normalized before matching)
REQUIRED = {
    "VVD": ["vvd"],
    "LANE": ["lane"],
    "CONT NO.": ["cont_nr", "cont no", "containerno", "container no", "container number", "cntr", "ctnno"],
    "FULL/EMPTY": ["fe_flg", "full/empty", "fullempty", "fe", "status"],
    "TYPE/SIZE": ["cont_tp_size_cd", "type/size", "typesize", "size", "container type", "ctype"],
    "POL": ["pol_cd", "pol", "port of loading", "load port"],
    "POD": ["pod_cd", "pod", "port of discharge", "discharge port", "dest port"],
    "CONT_Weight": ["cont_wt", "cont_weight", "cont weight", "weight", "gross weight", "contweight"],
    "AWK": ["awk_flg", "awk", "awkward"],
    "DG": ["dg_flg", "dg", "dangerous", "imdg"],
    "RF": ["rf_flg", "rf", "reefer"],
    "BB": ["bb_flg", "bb", "breakbulk", "break bulk"],
    "SLOT_OPR": ["slot_own_ptr_id", "slot_opr", "slot opr", "slot operator", "slotoperator"],
    "CONT_OPR": ["cont_opr_ptr_id", "cont_opr", "cont opr", "container operator", "contoperator", "operator"],
    "REVENUE MONTH": ["revenue_month", "revenue month", "revenuemonth", "rev month", "month"],
    # 源文件表头有拼写错误 TAGERT_PORT，这里两种都兼容
    "TARGET_PORT": ["tagert_port", "target_port", "target port", "targetport"],
}


def normalize(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


def build_header_map(headers):
    variants = {}
    for canon, alts in REQUIRED.items():
        variants[normalize(canon)] = canon
        for a in alts:
            variants[normalize(a)] = canon

    mapping = {}
    for idx, h in enumerate(headers):
        key = variants.get(normalize(h))
        if key and key not in mapping.values():
            mapping[idx] = key
    found = set(mapping.values())
    missing = [c for c in REQUIRED if c not in found]
    return mapping, missing


def convert(path, out_path, meta_path, gzip_out, limit=0):
    from openpyxl import load_workbook

    if gzip_out and not out_path.endswith(".gz"):
        out_path += ".gz"

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        headers = list(next(rows))
    except StopIteration:
        sys.exit("Excel 为空")

    mapping, missing = build_header_map(headers)
    if missing:
        print(f"[warn] 未匹配的列（将留空）: {missing}", file=sys.stderr)

    opener = gzip.open if gzip_out else open
    mode = "wt" if gzip_out else "w"

    count = 0
    with opener(out_path, mode, encoding="utf-8") as f:
        f.write("[")
        first = True
        for raw in rows:
            if raw is None or all(c is None for c in raw):
                continue
            rec = {}
            for idx, col in mapping.items():
                val = raw[idx] if idx < len(raw) else None
                if isinstance(val, datetime):
                    val = val.strftime("%Y-%m" if col == "REVENUE MONTH" else "%Y-%m-%d")
                rec[col] = "" if val is None else val
            f.write(("" if first else ",") + json.dumps(rec, ensure_ascii=False))
            first = False
            count += 1
            if limit and count >= limit:
                break
        f.write("]")

    meta = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "source": path,
        "rowCount": count,
        "compressed": gzip_out,
        "columns": list(REQUIRED.keys()),
        "missingColumns": missing,
    }
    with open(meta_path, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)

    size = 0
    try:
        import os
        size = os.path.getsize(out_path)
    except OSError:
        pass
    print(f"[ok] {count} 行 -> {out_path} ({size/1024/1024:.1f} MB, gzip={gzip_out})")
    print(f"[ok] 缺失列: {missing or '无'}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("input", help="源 Excel 路径 (.xlsx)")
    ap.add_argument("--out", default="data.json")
    ap.add_argument("--meta", default="meta.json")
    ap.add_argument("--gzip", action="store_true", help="输出压缩的 data.json.gz")
    ap.add_argument("--limit", type=int, default=0, help="仅转换前 N 行（用于测试/估算）")
    args = ap.parse_args()
    convert(args.input, args.out, args.meta, args.gzip, args.limit)
