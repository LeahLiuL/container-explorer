#!/usr/bin/env python3
"""Read-only inspection of a shipment Excel: print sheet, headers, sample rows,
and how well they map to the 16 required Container Explorer columns."""
import re
import sys
from openpyxl import load_workbook

TARGET = ["VVD","LANE","CONT NO.","FULL/EMPTY","TYPE/SIZE","POL","POD",
          "CONT_Weight","AWK","DG","RF","BB","SLOT_OPR","CONT_OPR",
          "REVENUE MONTH","TARGET_PORT"]

def norm(s): return re.sub(r"[^a-z0-9]", "", str(s).lower())

variants = {}
for c in TARGET:
    variants[norm(c)] = c
# extra aliases
alias = {
    "vvd":"VVD","vesselvoyage":"VVD","voyage":"VVD",
    "lane":"LANE","tradelane":"LANE","service":"LANE",
    "contno":"CONT NO.","containerno":"CONT NO.","containernumber":"CONT NO.","cntr":"CONT NO.",
    "fullempty":"FULL/EMPTY","fe":"FULL/EMPTY",
    "typesize":"TYPE/SIZE","containertype":"TYPE/SIZE","size":"TYPE/SIZE",
    "pol":"POL","portofloading":"POL","loadport":"POL",
    "pod":"POD","portofdischarge":"POD","dischargeport":"POD","destport":"POD",
    "contweight":"CONT_Weight","weight":"CONT_Weight","grossweight":"CONT_Weight",
    "awk":"AWK","awkward":"AWK","dg":"DG","dangerous":"DG","imdg":"DG",
    "rf":"RF","reefer":"RF","bb":"BB","breakbulk":"BB",
    "slotopr":"SLOT_OPR","slotoperator":"SLOT_OPR",
    "contopr":"CONT_OPR","containeroperator":"CONT_OPR","operator":"CONT_OPR",
    "revenuemonth":"REVENUE MONTH","revmonth":"REVENUE MONTH","month":"REVENUE MONTH",
    "targetport":"TARGET_PORT",
}
for k,v in alias.items(): variants[k]=v

path = sys.argv[1]
wb = load_workbook(path, read_only=True, data_only=True)
print("Sheets:", wb.sheetnames)
ws = wb.active
print("Active sheet:", ws.title, "| dims:", ws.calculate_dimension())
rows = ws.iter_rows(values_only=True)
try:
    headers = list(next(rows))
except StopIteration:
    print("Empty"); sys.exit()

print("\n=== Headers (col index : value) ===")
for i,h in enumerate(headers):
    print(f"  {i:>2} : {h!r}")

print("\n=== Mapping to required columns ===")
mapping = {}
for i,h in enumerate(headers):
    key = variants.get(norm(h))
    if key and key not in mapping.values():
        mapping[i] = key
        print(f"  [match] {h!r} -> {key}")
found = set(mapping.values())
for c in TARGET:
    if c not in found:
        print(f"  [MISSING] {c}")

print("\n=== First 3 data rows (mapped) ===")
for n in range(3):
    r = next(rows, None)
    if r is None: break
    d = {mapping[i]: (r[i] if i < len(r) else None) for i in mapping}
    for c in TARGET:
        print(f"    {c:>14}: {d.get(c)!r}")
    print("    " + "-"*30)
